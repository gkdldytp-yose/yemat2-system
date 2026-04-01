from __future__ import annotations

import csv
import io
import json
import math
import re
import uuid
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote
from xml.sax.saxutils import escape as xml_escape

try:
    import pandas as pd
    _PANDAS_IMPORT_ERROR = None
except ModuleNotFoundError as exc:
    pd = None
    _PANDAS_IMPORT_ERROR = exc
from flask import Blueprint, Response, flash, redirect, render_template, request, session, url_for

from core import (
    LOGISTICS_WORKPLACE,
    SHARED_WORKPLACE,
    WORKPLACES,
    admin_required,
    audit_log,
    get_db,
)

bp = Blueprint('imports', __name__)
PROJECT_ROOT = Path(__file__).resolve().parents[1]
UPLOAD_DIR = PROJECT_ROOT / 'uploads' / 'imports'
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

IMPORT_TYPE_LABELS = {
    'material_stock': '부자재 초기재고 임포트',
    'raw_material_stock': '원초 초기재고 임포트',
}

COMMON_ALIASES = {
    'name': ['품목명', '자재명', '상품명', '재료명', '명칭'],
    'code': ['품목코드', '자재코드', '코드', '품번'],
    'qty': ['수량', '재고', '현재고', '잔량'],
    'current_quantity': ['현재수량', '현재 재고'],
    'received_quantity': ['입고수량', '입고 수량', '총수량'],
    'received_date': ['입고일', '입고 날짜', '입고일자'],
    'manufacture_date': ['제조일', '제조일자'],
    'expiry_date': ['소비기한', '유통기한', '만료일'],
    'unit': ['단위'],
    'supplier_name': ['거래처', '거래처명', '매입처명', '업체', '업체명', '공급처', '제조사', '공급업체'],
    'workplace': ['작업장', '보관위치', '현장', '위치'],
    'category': ['분류', '카테고리'],
    'spec': ['규격', '스펙'],
    'lot': ['LOT', '로트', 'lot'],
    'supplier_lot': ['공급사LOT', '공급사 로트', 'supplier lot'],
    'unit_price': ['단가', '매입단가'],
    'ja_ho': ['자호'],
    'sheets_per_sok': ['속수', 'sok', '속'],
    'car_number': ['차량번호', '차량 번호'],
}

MATERIAL_FIELDS = [
    'code', 'name', 'category', 'spec', 'unit', 'qty', 'current_quantity',
    'received_quantity', 'received_date', 'manufacture_date', 'expiry_date',
    'supplier_name', 'workplace', 'lot', 'supplier_lot', 'unit_price',
]
RAW_FIELDS = [
    'code', 'name', 'qty', 'current_quantity', 'received_date', 'lot',
    'ja_ho', 'sheets_per_sok', 'car_number', 'workplace',
]

IMPORT_TEMPLATE_DEFINITIONS = {
    'products': {
        'filename': '상품관리_임포트_양식.xlsx',
        'sheets': [
            {
                'name': '상품_입력양식',
                'headers': ['상품코드', '상품명', '분류', '작업장', '매수', '절단', '한박스당원초', '비고'],
                'rows': [
                    ['A-KR-F002V1', '예맛 갯벌재래김 5g x 30봉', '갯벌김', '2동 신관 2층', 30, '5g', 2.8, '신규 등록 예시'],
                    ['A-EP-F002V1', '(대만) 갯벌재래김 5g x 30봉', '갯벌김', '2동 신관 2층', 30, '5g', 2.8, '작업장별 상품 관리'],
                ],
            },
            {
                'name': '상품_가이드',
                'headers': ['항목', '필수', '설명', '예시'],
                'rows': [
                    ['상품코드', '필수', '기존 시스템에서 식별하는 고유 상품코드', 'A-KR-F002V1'],
                    ['상품명', '필수', '상품관리에서 표시될 이름', '예맛 갯벌재래김 5g x 30봉'],
                    ['분류', '권장', '상품 분류값', '갯벌김'],
                    ['작업장', '필수', '생산/관리 작업장', '2동 신관 2층'],
                    ['매수', '권장', '상품 포장 기준 매수', '30'],
                    ['절단', '권장', '절단/규격 정보', '5g'],
                    ['한박스당원초', '권장', '1박스 생산 시 필요한 원초 사용량', '2.8'],
                ],
            },
        ],
    },
    'raw_materials': {
        'filename': '원초관리_임포트_양식.xlsx',
        'sheets': [
            {
                'name': '원초_입력양식',
                'headers': ['원초코드', '원초명', '속수', '입고일', 'LOT', '자호', '차량번호', '작업장', '현재고', '총재고', '사용수량', '비고'],
                'rows': [
                    ['A05', '곱창돌김', 100, '2026-03-19', 'A05-20260319-001', '1자호', '123가4567', '2동 신관 2층', 13394.0, 13394.0, 0.0, '초기 재고 이관 예시'],
                    ['A04', '감태김', 100, '2026-03-19', 'A04-20260319-001', '2자호', '', '1동 조미', 16083.6, 16083.6, 0.0, '차량번호 없으면 경고 처리'],
                ],
            },
            {
                'name': '원초_가이드',
                'headers': ['항목', '필수', '설명', '예시'],
                'rows': [
                    ['원초코드', '권장', '원초 식별 코드', 'A05'],
                    ['원초명', '필수', '원초명', '곱창돌김'],
                    ['속수', '필수', 'raw_materials.sheets_per_sok 로 반영', '100'],
                    ['입고일', '필수', 'YYYY-MM-DD 권장', '2026-03-19'],
                    ['LOT', '권장', '기존 lot가 있으면 그대로 입력', 'A05-20260319-001'],
                    ['자호', '권장', '원초 추적용 자호', '1자호'],
                    ['차량번호', '권장', '입고 차량 정보', '123가4567'],
                    ['작업장', '필수', '원초를 보유한 작업장', '2동 신관 2층'],
                    ['현재고', '필수', '현재 남은 수량', '13394.0'],
                    ['총재고', '권장', '초기 이식 시 현재고와 동일하게 맞추는 것을 권장', '13394.0'],
                    ['사용수량', '권장', '이미 사용 이력이 있으면 입력, 없으면 0', '0'],
                ],
            },
        ],
    },
    'materials': {
        'filename': '부자재관리_임포트_양식.xlsx',
        'sheets': [
            {
                'name': '부자재_입력양식',
                'headers': ['자재코드', '자재명', '분류', '단위', '공급처', '보유위치', '현재수량', '입고수량', '입고일', '제조일', '소비기한', '단가', '비고'],
                'rows': [
                    ['P02A005', '국내 두번구운곱창돌김식탁17g 10봉 외포', '외포', '롤', '영진물산', '물류', 16.0, 36.0, '2026-03-17', '', '', 0, '물류 Hub 보유 재고'],
                    ['P02A005', '국내 두번구운곱창돌김식탁17g 10봉 외포', '외포', '롤', '영진물산', '2동 신관 2층', 11.0, 0.0, '2026-03-17', '', '', 0, '같은 코드라도 작업장 보유분은 별도 행'],
                    ['O01', '참기름_조미', '기름', 'kg', '성창실업', '물류', 6600.0, 6600.0, '2026-03-19', '', '', 0, '작업장에 없고 물류만 있는 기본 예시'],
                ],
            },
            {
                'name': '부자재_가이드',
                'headers': ['항목', '필수', '설명', '예시'],
                'rows': [
                    ['자재코드', '필수 권장', 'materials.code 우선 매칭에 사용', 'P02A005'],
                    ['자재명', '필수', 'materials.name 매칭 또는 신규 생성 후보 판단에 사용', '국내 두번구운곱창돌김식탁17g 10봉 외포'],
                    ['분류', '권장', '기름/소금/내포/외포/박스/실리카 등', '외포'],
                    ['단위', '권장', '롤/kg/ea 등', '롤'],
                    ['공급처', '권장', 'suppliers.name 매칭', '영진물산'],
                    ['작업장별 시트', '필수 입력', '각 작업장 시트의 재고 컬럼에 해당 작업장 실재고를 입력합니다.', '2동 신관 2층 시트 F열'],
                    ['물류 재고', '필수 입력', '물류 Hub 재고는 종합 시트의 물류 재고 컬럼에만 입력합니다.', '부자재_종합 K열'],
                    ['총재고', '자동 계산', '각 작업장 재고 + 물류 재고 합계가 자동 계산됩니다.', '자동'],
                    ['입고수량', '권장', '초기 로트 총 입고량. 없으면 현재수량 기준으로 처리 가능', '36.0'],
                    ['입고일', '필수 권장', '시스템 LOT 자동생성 기준. 사용자가 LOT를 직접 적지 않습니다.', '2026-03-17'],
                    ['제조일', '권장', '없어도 임포트는 되지만 정보미흡 상태로 등록됩니다. 생산 사용 시 보완 입력할 수 있습니다.', '2026-03-01'],
                    ['소비기한', '기름/소금 권장', '기름/소금류는 소비기한이 없으면 정보미흡 상태로 등록됩니다. 그 외 부자재는 비워둘 수 있습니다.', '2027-03-01'],
                    ['단가', '선택', 'material_lots.unit_price 및 materials.unit_price 후보', '0'],
                    ['중요 규칙', '반드시 확인', '물류는 Hub이며 작업장 재고와 분리 관리합니다. 물류 수량은 종합 시트, 작업장 수량은 각 작업장 시트에 나눠 입력합니다.', '물류 16 / 작업장 11 분리 입력'],
                    ['LOT 처리', '자동', 'LOT는 시스템이 코드+입고일+순번 기준으로 생성합니다. 엑셀 입력 컬럼이 없습니다.', 'P02A005-20260317-001'],
                ],
            },
        ],
    },
}

TEMPLATE_GUIDANCE_ROWS = {
    'materials': [
        ['작성방법', '안내', '같은 자재코드라도 보유위치가 다르면 행을 분리합니다.', '', '', '', '', '', '', '', '', '', ''],
        ['작성방법', '안내', 'LOT는 시스템이 코드+입고일+순번으로 자동 생성합니다.', '', '', '', '', '', '', '', '', '', ''],
        ['작성방법', '안내', '제조일/소비기한이 비어 있으면 업로드 후 WARNING으로 표시되며 보완 입력할 수 있습니다.', '', '', '', '', '', '', '', '', '', ''],
    ],
}


def _is_import_runtime_ready() -> bool:
    return pd is not None


def _missing_runtime_message() -> str:
    return "엑셀 임포트 기능을 사용하려면 pandas가 설치되어 있어야 합니다."


def _xlsx_escape_text(value: Any) -> str:
    text = '' if value is None else str(value)
    text = re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f]', '', text)
    return xml_escape(text)


def _xlsx_column_name(index: int) -> str:
    result = ''
    current = int(index)
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        result = chr(65 + remainder) + result
    return result or 'A'


def _build_xlsx_workbook(sheets: list[dict[str, Any]]) -> io.BytesIO:
    buffer = io.BytesIO()

    def _cell_xml(row_idx: int, col_idx: int, value: Any) -> str:
        cell_ref = f"{_xlsx_column_name(col_idx)}{row_idx}"
        if isinstance(value, dict) and value.get('formula'):
            formula = _xlsx_escape_text(value.get('formula'))
            return f'<c r="{cell_ref}"><f>{formula}</f></c>'
        if value is None or value == '':
            return f'<c r="{cell_ref}" t="inlineStr"><is><t></t></is></c>'
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return f'<c r="{cell_ref}"><v>{value}</v></c>'
        return f'<c r="{cell_ref}" t="inlineStr"><is><t>{_xlsx_escape_text(value)}</t></is></c>'

    workbook_sheet_xml = []
    workbook_rel_xml = []
    content_override_xml = []

    with zipfile.ZipFile(buffer, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        for idx, sheet in enumerate(sheets, start=1):
            headers = sheet.get('headers', [])
            rows = sheet.get('rows', [])
            all_rows = [headers] + rows
            max_cols = max((len(r) for r in all_rows), default=0)
            cols_xml = ''.join(
                f'<col min="{col_idx}" max="{col_idx}" width="20" customWidth="1"/>'
                for col_idx in range(1, max_cols + 1)
            )
            row_xml = []
            for row_idx, row in enumerate(all_rows, start=1):
                cell_xml = ''.join(
                    _cell_xml(row_idx, col_idx, value)
                    for col_idx, value in enumerate(row, start=1)
                )
                row_xml.append(f'<row r="{row_idx}">{cell_xml}</row>')

            sheet_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                f'<cols>{cols_xml}</cols>'
                f'<sheetData>{"".join(row_xml)}</sheetData>'
                '</worksheet>'
            )
            zf.writestr(f'xl/worksheets/sheet{idx}.xml', sheet_xml)
            workbook_sheet_xml.append(
                f'<sheet name="{_xlsx_escape_text(sheet.get("name") or f"Sheet{idx}")}" sheetId="{idx}" r:id="rId{idx}"/>'
            )
            workbook_rel_xml.append(
                '<Relationship Id="rId{0}" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
                'Target="worksheets/sheet{0}.xml"/>'.format(idx)
            )
            content_override_xml.append(
                '<Override PartName="/xl/worksheets/sheet{0}.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'.format(idx)
            )

        workbook_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            f'<sheets>{"".join(workbook_sheet_xml)}</sheets>'
            '<calcPr calcId="171027" fullCalcOnLoad="1"/>'
            '</workbook>'
        )
        workbook_rels_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            f'{"".join(workbook_rel_xml)}'
            f'<Relationship Id="rId{len(sheets) + 1}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" '
            'Target="styles.xml"/>'
            '</Relationships>'
        )
        root_rels_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
            'Target="xl/workbook.xml"/>'
            '</Relationships>'
        )
        content_types_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            f'{"".join(content_override_xml)}'
            '<Override PartName="/xl/styles.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
            '</Types>'
        )
        styles_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<fonts count="1"><font><sz val="11"/><name val="Malgun Gothic"/></font></fonts>'
            '<fills count="1"><fill><patternFill patternType="none"/></fill></fills>'
            '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
            '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
            '<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>'
            '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
            '</styleSheet>'
        )
        zf.writestr('[Content_Types].xml', content_types_xml)
        zf.writestr('_rels/.rels', root_rels_xml)
        zf.writestr('xl/workbook.xml', workbook_xml)
        zf.writestr('xl/_rels/workbook.xml.rels', workbook_rels_xml)
        zf.writestr('xl/styles.xml', styles_xml)

    buffer.seek(0)
    return buffer


def _build_template_response(template_key: str) -> Response:
    definition = _get_import_template_definition(template_key)
    if not definition:
        raise KeyError(template_key)
    workbook = _build_xlsx_workbook(definition['sheets'])
    filename = definition['filename']
    quoted_name = quote(filename)
    response = Response(
        workbook.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response.headers['Content-Disposition'] = (
        f"attachment; filename*=UTF-8''{quoted_name}"
    )
    return response


def _get_import_template_definition(template_key: str) -> dict[str, Any] | None:
    definition = IMPORT_TEMPLATE_DEFINITIONS.get(template_key)
    if not definition:
        return None
    if template_key != 'materials':
        return definition

    conn = get_db()
    cursor = conn.cursor()
    try:
        material_rows = cursor.execute(
            '''
            SELECT code, name, category, unit, workplace, supplier_id
            FROM materials
            ORDER BY category, name
            '''
        ).fetchall()
        supplier_map = {
            int(row['id']): row['name']
            for row in cursor.execute('SELECT id, name FROM suppliers').fetchall()
        }
        workplace_sheets = list(WORKPLACES)
        workplace_column_order = list(WORKPLACES)
        material_master_rows = []
        for row in material_rows:
            supplier_name = supplier_map.get(int(row['supplier_id'])) if row['supplier_id'] else ''
            material_master_rows.append([
                row['code'] or '',
                row['name'] or '',
                row['category'] or '',
                _normalize_material_unit(row['unit']),
                supplier_name or '',
                row['workplace'] or '',
            ])
        summary_headers = ['자재코드', '자재명', '분류', '단위', '공급처', '기본작업장']
        summary_headers.extend(workplace_column_order)
        summary_headers.extend(['물류 재고', '총재고', '입고수량', '입고일', '제조일(권장)', '소비기한(기름/소금 권장)', '단가', '비고'])

        summary_rows = []
        data_start_row = 2
        workplace_col_map = {}
        for idx, workplace in enumerate(workplace_column_order, start=7):
            workplace_col_map[workplace] = _xlsx_column_name(idx)
        logistics_col = _xlsx_column_name(7 + len(workplace_column_order))
        total_col = _xlsx_column_name(8 + len(workplace_column_order))
        receipt_col = _xlsx_column_name(9 + len(workplace_column_order))
        received_date_col = _xlsx_column_name(10 + len(workplace_column_order))
        manufacture_col = _xlsx_column_name(11 + len(workplace_column_order))
        expiry_col = _xlsx_column_name(12 + len(workplace_column_order))
        unit_price_col = _xlsx_column_name(13 + len(workplace_column_order))
        note_col = _xlsx_column_name(14 + len(workplace_column_order))

        workplace_sheet_rows: dict[str, list[list[Any]]] = {workplace: [] for workplace in workplace_sheets}
        for row_index, master_row in enumerate(material_master_rows, start=data_start_row):
            summary_row = list(master_row)
            for workplace in workplace_column_order:
                summary_row.append({'formula': f"'{workplace}'!F{row_index}"})
            summary_row.append('')
            total_formula = '+'.join(
                [f'{workplace_col_map[wp]}{row_index}' for wp in workplace_column_order] + [f'{logistics_col}{row_index}']
            )
            summary_row.append({'formula': total_formula})
            summary_row.extend(['', '', '', '', '', ''])
            summary_rows.append(summary_row)

            for workplace in workplace_sheets:
                workplace_sheet_rows[workplace].append([
                    master_row[0],
                    master_row[1],
                    master_row[2],
                    master_row[3],
                    master_row[4],
                    '',
                    {'formula': f"'부자재_종합'!{receipt_col}{row_index}"},
                    {'formula': f"'부자재_종합'!{received_date_col}{row_index}"},
                    {'formula': f"'부자재_종합'!{manufacture_col}{row_index}"},
                    {'formula': f"'부자재_종합'!{expiry_col}{row_index}"},
                    {'formula': f"'부자재_종합'!{unit_price_col}{row_index}"},
                    {'formula': f"'부자재_종합'!{note_col}{row_index}"},
                ])

        sheets = [{
            'name': '부자재_종합',
            'headers': summary_headers,
            'rows': summary_rows,
        }]
        for workplace in workplace_sheets:
            sheets.append({
                'name': workplace,
                'headers': ['자재코드', '자재명', '분류', '단위', '공급처', f'{workplace} 재고', '입고수량', '입고일', '제조일(권장)', '소비기한(기름/소금 권장)', '단가', '비고'],
                'rows': workplace_sheet_rows[workplace],
            })
        for sheet in definition['sheets']:
            if sheet['name'] == '부자재_가이드':
                sheets.append({
                    'name': sheet['name'],
                    'headers': list(sheet['headers']),
                    'rows': [list(item) for item in sheet['rows']],
                })
        return {
            'filename': definition['filename'],
            'sheets': sheets,
        }
    finally:
        conn.close()


def _json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=_json_default)


def _json_default(value: Any):
    if hasattr(value, 'isoformat'):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value)


def _json_loads(value: str | None, default: Any):
    if not value:
        return default
    try:
        return json.loads(value)
    except Exception:
        return default


def _normalize_header(value: Any) -> str:
    text = str(value or '').strip().lower()
    text = text.replace('\n', ' ')
    text = re.sub(r'\s+', '', text)
    return text


def _normalize_text(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and math.isnan(value):
        return ''
    return str(value).strip()


def _clean_optional_text(value: Any) -> str | None:
    text = _normalize_text(value)
    return text or None


def _parse_float(value: Any) -> float | None:
    text = _normalize_text(value)
    if not text:
        return None
    text = text.replace(',', '')
    try:
        return float(text)
    except Exception:
        return None


def _parse_int(value: Any) -> int | None:
    parsed = _parse_float(value)
    if parsed is None:
        return None
    try:
        return int(parsed)
    except Exception:
        return None


def _round_qty(value: Any) -> float:
    try:
        return round(float(value or 0), 1)
    except Exception:
        return 0.0


def _parse_date(value: Any) -> tuple[str | None, str | None]:
    raw = _normalize_text(value)
    if not raw:
        return None, None
    try:
        parsed = pd.to_datetime(raw, errors='raise')
        return parsed.strftime('%Y-%m-%d'), None
    except Exception:
        return None, f'날짜 형식 오류: {raw}'


def _normalize_material_unit(value: Any) -> str:
    text = _normalize_text(value)
    if not text:
        return ''
    lowered = text.lower().replace(' ', '')
    replacements = {
        '1kg': 'kg',
        'kg': 'kg',
        '1g': 'g',
        'g': 'g',
        '1l': 'L',
        'l': 'L',
        '1ml': 'ml',
        'ml': 'ml',
        'ea': 'ea',
        '개': '개',
        '롤': '롤',
        '박스': '박스',
    }
    return replacements.get(lowered, text)


def _normalize_import_workplace_label(value: Any) -> str | None:
    text = _normalize_text(value)
    if not text:
        return None
    key = _normalize_header(text)
    if '1동자반' in key:
        return '1동 자반'
    if '1동조미' in key or '조미김창고' in key and '1동' in key:
        return '1동 조미'
    if '2동1f' in key or '2동1층' in key or '2동신관1층' in key:
        return '2동 신관 1층'
    if '2동2f' in key or '2동2층' in key or '2동신관2층' in key:
        return '2동 신관 2층'
    if '물류' in key or '로지스' in key or 'hub' in key:
        return LOGISTICS_WORKPLACE
    return None


def _load_material_ledger_master_sheet(file_path: Path) -> dict[str, dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, data_only=True)
    if '품번코드마스터' not in workbook.sheetnames:
        return {}

    ws = workbook['품번코드마스터']
    master: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        code = _normalize_text(row[0] if len(row) > 0 else None)
        if not code:
            continue
        master[code] = {
            'code': code,
            'category': _normalize_text(row[1] if len(row) > 1 else None),
            'supplier_name': _normalize_text(row[2] if len(row) > 2 else None),
            'description': _normalize_text(row[3] if len(row) > 3 else None),
            'name': _normalize_text(row[4] if len(row) > 4 else None),
            'unit': _normalize_material_unit(row[5] if len(row) > 5 else None),
            'warehouse_name': _normalize_text(row[6] if len(row) > 6 else None),
            'warehouse_code': _normalize_text(row[7] if len(row) > 7 else None),
            'location_code': _normalize_text(row[8] if len(row) > 8 else None),
            'unit_price': _parse_float(row[9] if len(row) > 9 else None) or 0.0,
        }
    return master


def _should_include_material_ledger_row(code: str, name: str, category_text: str, master_row: dict[str, Any] | None) -> bool:
    normalized_category = _normalize_text(category_text)
    if '부재료' in normalized_category:
        return True

    guessed = _guess_material_category(name)
    if guessed in {'기름', '소금', '실리카', '내포', '외포', '박스', '트레이', '뚜껑', '각대'}:
        return True

    master_category = _normalize_text((master_row or {}).get('category'))
    if any(keyword in master_category for keyword in ['부재료', '기름', '소금', '실리카', '내포', '외포', '박스', '트레이', '뚜껑', '각대']):
        return True

    code_prefix = _normalize_text(code)[:1]
    if code_prefix in {'O', 'B', 'S', 'T', 'P', 'Z'}:
        return True

    return False


def _parse_material_ledger_workbook(cursor, batch_id: int, file_path: Path, seq_cache: dict[tuple[str, str], int]) -> bool:
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, data_only=True)
    if '원,부재료 수불부' not in workbook.sheetnames:
        return False

    ws = workbook['원,부재료 수불부']
    if _normalize_text(ws.cell(row=5, column=2).value) != '원,부재료코드':
        return False

    master_lookup = _load_material_ledger_master_sheet(file_path)
    workplace_columns = []
    for col_idx in range(21, ws.max_column + 1):
        workplace = _normalize_import_workplace_label(ws.cell(row=5, column=col_idx).value)
        if workplace:
            workplace_columns.append((col_idx, workplace))

    if not workplace_columns:
        return False

    mapping_summary = [{
        'sheet_name': ws.title,
        'mapping': {
            'code': '원,부재료코드',
            'name': '품명',
            'supplier_name': '거래처명(매입처명)',
            'unit': '관리단위',
            'qty': '실재고/작업장별 수량',
            'workplace': '작업장별 재고 컬럼',
        },
        'headers': [str(ws.cell(row=5, column=col).value or '') for col in range(1, ws.max_column + 1)],
        'parser': 'material_ledger',
    }]

    for excel_row_no in range(7, ws.max_row + 1):
        code = _normalize_text(ws.cell(row=excel_row_no, column=2).value)
        if not code:
            continue

        category_text = _normalize_text(ws.cell(row=excel_row_no, column=5).value)
        if not _should_include_material_ledger_row(code, _normalize_text(ws.cell(row=excel_row_no, column=3).value), category_text, master_lookup.get(code)):
            continue

        master_row = master_lookup.get(code, {})
        name = _normalize_text(ws.cell(row=excel_row_no, column=3).value) or master_row.get('name') or ''
        supplier_name = _normalize_text(ws.cell(row=excel_row_no, column=4).value) or master_row.get('supplier_name') or ''
        unit = _normalize_material_unit(ws.cell(row=excel_row_no, column=6).value) or master_row.get('unit') or ''
        material_category = _normalize_text(master_row.get('category')) or _guess_material_category(name)
        unit_price = _parse_float(ws.cell(row=excel_row_no, column=12).value)
        if unit_price is None:
            unit_price = float(master_row.get('unit_price') or 0.0)
        actual_stock = _parse_float(ws.cell(row=excel_row_no, column=20).value) or 0.0

        generated_any = False
        workplace_quantities = []
        for col_idx, workplace in workplace_columns:
            qty = _parse_float(ws.cell(row=excel_row_no, column=col_idx).value)
            if qty is None or qty <= 0:
                continue
            workplace_quantities.append((workplace, qty))
        workplace_total = sum(qty for _, qty in workplace_quantities)

        for workplace, qty in workplace_quantities:
            generated_any = True
            raw_row = {
                'source_parser': 'material_ledger',
                'excel_row_no': excel_row_no,
                'code': code,
                'name': name,
                'supplier_name': supplier_name,
                'category': material_category,
                'unit': unit,
                'current_quantity': qty,
                'qty': qty,
                'received_quantity': qty,
                'received_date': datetime.now().strftime('%Y-%m-%d'),
                'workplace': workplace,
                'unit_price': unit_price,
                'source_actual_stock': actual_stock,
            }
            cursor.execute(
                '''
                INSERT INTO import_raw_rows (batch_id, sheet_name, row_no, raw_json)
                VALUES (?, ?, ?, ?)
                ''',
                (batch_id, ws.title, excel_row_no, _json_dumps(raw_row)),
            )
            parsed = _validate_material_parsed(cursor, raw_row, seq_cache)
            if actual_stock and abs(workplace_total - actual_stock) > 0.001:
                warning = parsed.get('warning_message') or ''
                extra = f'작업장 합계({workplace_total})와 실재고({actual_stock}) 차이를 확인해주세요.'
                parsed['warning_message'] = f'{warning} | {extra}'.strip(' |')
                parsed['status'] = 'WARNING' if parsed['status'] == 'OK' else parsed['status']
            _save_parsed_row(cursor, batch_id, ws.title, excel_row_no, parsed)

        if not generated_any and actual_stock > 0:
            raw_row = {
                'source_parser': 'material_ledger',
                'excel_row_no': excel_row_no,
                'code': code,
                'name': name,
                'supplier_name': supplier_name,
                'category': material_category,
                'unit': unit,
                'current_quantity': actual_stock,
                'qty': actual_stock,
                'received_quantity': actual_stock,
                'received_date': datetime.now().strftime('%Y-%m-%d'),
                'workplace': LOGISTICS_WORKPLACE,
                'unit_price': unit_price,
            }
            cursor.execute(
                '''
                INSERT INTO import_raw_rows (batch_id, sheet_name, row_no, raw_json)
                VALUES (?, ?, ?, ?)
                ''',
                (batch_id, ws.title, excel_row_no, _json_dumps(raw_row)),
            )
            parsed = _validate_material_parsed(cursor, raw_row, seq_cache)
            warning = parsed.get('warning_message') or ''
            extra = '작업장 배분값이 없어 이 행의 실재고를 물류 재고로 반영했습니다.'
            parsed['warning_message'] = f'{warning} | {extra}'.strip(' |')
            parsed['status'] = 'WARNING' if parsed['status'] == 'OK' else parsed['status']
            _save_parsed_row(cursor, batch_id, ws.title, excel_row_no, parsed)

    stats = _recompute_batch_stats(cursor, batch_id)
    cursor.execute(
        '''
        UPDATE import_batches
        SET column_mapping_json = ?, status = 'validated'
        WHERE id = ?
        ''',
        (_json_dumps(mapping_summary), batch_id),
    )
    audit_log(cursor.connection, 'validate', 'import_batches', batch_id, stats)
    return True


def _material_workplace_choices() -> list[str]:
    return [*WORKPLACES, SHARED_WORKPLACE, LOGISTICS_WORKPLACE]


def _raw_workplace_choices() -> list[str]:
    return [*WORKPLACES, SHARED_WORKPLACE]


def _detect_column_mapping(headers: list[str], import_type: str) -> dict[str, str]:
    target_fields = MATERIAL_FIELDS if import_type == 'material_stock' else RAW_FIELDS
    normalized_headers = {_normalize_header(col): str(col) for col in headers}
    mapping: dict[str, str] = {}
    for field in target_fields:
        aliases = COMMON_ALIASES.get(field, [])
        for alias in aliases:
            candidate = normalized_headers.get(_normalize_header(alias))
            if candidate:
                mapping[field] = candidate
                break
    return mapping


def _extract_mapped_value(row: dict[str, Any], mapping: dict[str, str], field: str) -> Any:
    source_col = mapping.get(field)
    if not source_col:
        return None
    return row.get(source_col)


def _load_material_options(cursor) -> list[dict[str, Any]]:
    rows = cursor.execute(
        '''
        SELECT id, code, name, category, unit, workplace
        FROM materials
        ORDER BY name
        '''
    ).fetchall()
    return [dict(row) for row in rows]


def _load_raw_material_options(cursor) -> list[dict[str, Any]]:
    rows = cursor.execute(
        '''
        SELECT id, code, name, lot, ja_ho, workplace, current_stock
        FROM raw_materials
        ORDER BY name, receiving_date DESC, id DESC
        '''
    ).fetchall()
    return [dict(row) for row in rows]


def _load_supplier_lookup(cursor) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    by_name = {}
    by_code = {}
    rows = cursor.execute('SELECT id, code, name FROM suppliers ORDER BY name').fetchall()
    for row in rows:
        item = dict(row)
        if item['name']:
            by_name[_normalize_header(item['name'])] = item
        if item['code']:
            by_code[_normalize_header(item['code'])] = item
    return by_name, by_code


def _load_material_lookup(cursor):
    by_code = {}
    by_name = {}
    rows = cursor.execute(
        '''
        SELECT id, code, name, category, unit, supplier_id, workplace, spec, unit_price
        FROM materials
        ORDER BY id
        '''
    ).fetchall()
    for row in rows:
        item = dict(row)
        code = _normalize_header(item.get('code'))
        name = _normalize_header(item.get('name'))
        if code:
            by_code[code] = item
        if name:
            by_name[name] = item
    mappings = cursor.execute('SELECT source_text, material_id FROM material_name_mappings').fetchall()
    name_map = {}
    for row in mappings:
        if row['source_text']:
            name_map[_normalize_header(row['source_text'])] = int(row['material_id'])
    return by_code, by_name, name_map


def _load_raw_lookup(cursor):
    by_id = {}
    by_key = {}
    by_name = {}
    rows = cursor.execute(
        '''
        SELECT id, code, name, lot, ja_ho, receiving_date, workplace, current_stock, total_stock, used_quantity, sheets_per_sok, car_number
        FROM raw_materials
        ORDER BY id
        '''
    ).fetchall()
    for row in rows:
        item = dict(row)
        by_id[int(item['id'])] = item
        code = _normalize_text(item.get('code'))
        lot = _normalize_text(item.get('lot'))
        ja_ho = _normalize_text(item.get('ja_ho'))
        receiving_date = _normalize_text(item.get('receiving_date'))
        if code and lot:
            by_key[('code_lot', code, lot)] = item
        if code and receiving_date and ja_ho:
            by_key[('code_date_ja', code, receiving_date, ja_ho)] = item
        name_key = _normalize_header(item.get('name'))
        if name_key:
            by_name.setdefault(name_key, []).append(item)
    mappings = cursor.execute('SELECT source_text, raw_material_code, raw_material_name FROM raw_material_name_mappings').fetchall()
    name_map = {}
    for row in mappings:
        source = _normalize_header(row['source_text'])
        if source:
            name_map[source] = {
                'code': _normalize_text(row['raw_material_code']),
                'name': _normalize_text(row['raw_material_name']),
            }
    return by_id, by_key, by_name, name_map


def _make_batch_upload_name(batch_id: int, original_name: str) -> str:
    suffix = Path(original_name or '').suffix.lower() or '.xlsx'
    token = uuid.uuid4().hex[:10]
    return f'batch_{batch_id}_{token}{suffix}'


def _guess_material_category(name: str) -> str:
    text = _normalize_text(name)
    if '기름' in text or '해바라기유' in text or '참기름' in text:
        return '기름'
    if '소금' in text:
        return '소금'
    if '내포' in text:
        return '내포'
    if '외포' in text:
        return '외포'
    if '박스' in text:
        return '박스'
    if '실리카' in text:
        return '실리카'
    if '트레이' in text:
        return '트레이'
    if '뚜껑' in text:
        return '뚜껑'
    if '각대' in text:
        return '각대'
    return '기타'


def _material_requires_expiry(category: Any) -> bool:
    return _normalize_text(category) == '기름'


def _generate_material_lot(cursor, code: str, received_date: str | None, seq_cache: dict[tuple[str, str], int]) -> tuple[str, int]:
    date_token = (received_date or datetime.now().strftime('%Y-%m-%d')).replace('-', '')
    code_token = _normalize_text(code) or 'MAT'
    cache_key = (code_token, date_token)
    if cache_key not in seq_cache:
        prefix = f'{code_token}-{date_token}-'
        row = cursor.execute(
            '''
            SELECT COALESCE(MAX(lot_seq), 0) AS max_seq
            FROM material_lots
            WHERE lot LIKE ?
            ''',
            (f'{prefix}%',),
        ).fetchone()
        seq_cache[cache_key] = int((row['max_seq'] if row else 0) or 0)
    seq_cache[cache_key] += 1
    seq = seq_cache[cache_key]
    return f'{code_token}-{date_token}-{seq:03d}', seq

def _validate_material_parsed(cursor, parsed: dict[str, Any], seq_cache: dict[tuple[str, str], int]) -> dict[str, Any]:
    errors = []
    warnings = []
    by_code, by_name, name_map = _load_material_lookup(cursor)
    supplier_by_name, supplier_by_code = _load_supplier_lookup(cursor)

    code = _normalize_text(parsed.get('code'))
    name = _normalize_text(parsed.get('name'))
    if not code and not name:
        errors.append('code와 name이 모두 없습니다.')

    qty = _parse_float(parsed.get('current_quantity'))
    if qty is None:
        qty = _parse_float(parsed.get('qty'))
    if qty is None:
        errors.append('수량 숫자 변환 실패')
        qty = 0.0
    received_qty = _parse_float(parsed.get('received_quantity'))
    if received_qty is None:
        received_qty = qty
    qty = _round_qty(qty)
    received_qty = _round_qty(received_qty)

    received_date, received_date_error = _parse_date(parsed.get('received_date'))
    manufacture_date, manufacture_date_error = _parse_date(parsed.get('manufacture_date'))
    expiry_date, expiry_date_error = _parse_date(parsed.get('expiry_date'))
    if received_date_error:
        errors.append(received_date_error)
    if manufacture_date_error:
        errors.append(manufacture_date_error)
    if expiry_date_error:
        errors.append(expiry_date_error)

    matched_material = None
    matched_material_id = parsed.get('matched_material_id')
    if matched_material_id:
        row = cursor.execute('SELECT * FROM materials WHERE id = ?', (matched_material_id,)).fetchone()
        if row:
            matched_material = dict(row)
    if not matched_material and code:
        matched_material = by_code.get(_normalize_header(code))
    if not matched_material and name:
        mapped_id = name_map.get(_normalize_header(name))
        if mapped_id:
            row = cursor.execute('SELECT * FROM materials WHERE id = ?', (mapped_id,)).fetchone()
            if row:
                matched_material = dict(row)
        if not matched_material:
            matched_material = by_name.get(_normalize_header(name))
            if matched_material:
                warnings.append('code 없이 name 매핑만 된 경우')

    supplier_name = _normalize_text(parsed.get('supplier_name'))
    supplier_id = parsed.get('supplier_id')
    supplier = None
    if supplier_id:
        row = cursor.execute('SELECT id, name, code FROM suppliers WHERE id = ?', (supplier_id,)).fetchone()
        if row:
            supplier = dict(row)
    if not supplier and supplier_name:
        supplier = supplier_by_name.get(_normalize_header(supplier_name)) or supplier_by_code.get(_normalize_header(supplier_name))
    if not supplier_name:
        warnings.append('공급처 없음')

    lot = _normalize_text(parsed.get('lot'))
    lot_seq = _parse_int(parsed.get('lot_seq')) or 1
    if not lot and code:
        lot, lot_seq = _generate_material_lot(cursor, code, received_date, seq_cache)
    elif not lot:
        errors.append('lot 자동 생성에 필요한 code가 없습니다.')
    if lot:
        duplicate = cursor.execute('SELECT id, material_id FROM material_lots WHERE lot = ?', (lot,)).fetchone()
        if duplicate and matched_material and int(duplicate['material_id']) != int(matched_material['id']):
            errors.append('lot 중복 충돌')

    if not matched_material:
        warnings.append('기존 materials 미매칭, 신규 생성 후보')

    workplace = _normalize_text(parsed.get('workplace')) or (matched_material.get('workplace') if matched_material else '1동 조미')
    if workplace not in _material_workplace_choices():
        warnings.append('작업장 값이 기본값으로 보정됩니다.')
        workplace = '1동 조미'

    category = _normalize_text(parsed.get('category')) or (matched_material.get('category') if matched_material else _guess_material_category(name))
    unit = _normalize_material_unit(parsed.get('unit')) or (matched_material.get('unit') if matched_material else '')
    spec = _normalize_text(parsed.get('spec')) or (matched_material.get('spec') if matched_material else '')
    unit_price = _parse_float(parsed.get('unit_price'))
    if unit_price is None:
        unit_price = float(matched_material.get('unit_price') or 0) if matched_material else 0.0

    status = 'ERROR' if errors else ('WARNING' if warnings else 'OK')
    return {
        **parsed,
        'target_type': 'material',
        'matched_material_id': int(matched_material['id']) if matched_material else None,
        'supplier_id': int(supplier['id']) if supplier else None,
        'supplier_name': supplier_name or (supplier['name'] if supplier else None),
        'code': code or (matched_material.get('code') if matched_material else ''),
        'name': name or (matched_material.get('name') if matched_material else ''),
        'category': category,
        'spec': spec,
        'unit': unit,
        'qty': qty,
        'received_quantity': received_qty,
        'current_quantity': qty,
        'received_date': received_date or datetime.now().strftime('%Y-%m-%d'),
        'manufacture_date': manufacture_date,
        'expiry_date': expiry_date,
        'lot': lot,
        'lot_seq': lot_seq,
        'supplier_lot': _clean_optional_text(parsed.get('supplier_lot')),
        'workplace': workplace,
        'unit_price': unit_price,
        'status': status,
        'error_message': ' | '.join(errors),
        'warning_message': ' | '.join(warnings),
    }


def _validate_raw_parsed(cursor, parsed: dict[str, Any]) -> dict[str, Any]:
    errors = []
    warnings = []
    by_id, by_key, by_name, name_map = _load_raw_lookup(cursor)

    code = _normalize_text(parsed.get('code'))
    name = _normalize_text(parsed.get('name'))
    lot = _normalize_text(parsed.get('lot'))
    ja_ho = _normalize_text(parsed.get('ja_ho'))
    if not name:
        errors.append('name 없음')
    sheets_per_sok = _parse_int(parsed.get('sheets_per_sok'))
    if sheets_per_sok is None:
        errors.append('sheets_per_sok 없음 또는 숫자 아님')
        sheets_per_sok = 0
    qty = _parse_float(parsed.get('current_quantity'))
    if qty is None:
        qty = _parse_float(parsed.get('qty'))
    if qty is None:
        errors.append('current_stock 숫자 아님')
        qty = 0.0
    qty = _round_qty(qty)
    received_date, received_date_error = _parse_date(parsed.get('received_date'))
    if received_date_error:
        errors.append(received_date_error)
    if not code and not lot and not ja_ho:
        errors.append('code/lot 식별 불가')
    if not ja_ho:
        warnings.append('ja_ho 없음')
    if not _normalize_text(parsed.get('car_number')):
        warnings.append('car_number 없음')

    matched = None
    matched_raw_material_id = parsed.get('matched_raw_material_id')
    if matched_raw_material_id:
        matched = by_id.get(int(matched_raw_material_id))
    if not matched and code and lot:
        matched = by_key.get(('code_lot', code, lot))
    if not matched and code and received_date and ja_ho:
        matched = by_key.get(('code_date_ja', code, received_date, ja_ho))
    if not matched and name:
        mapped = name_map.get(_normalize_header(name))
        if mapped:
            for item in by_id.values():
                if _normalize_text(item.get('code')) == mapped.get('code') and _normalize_header(item.get('name')) == _normalize_header(mapped.get('name')):
                    matched = item
                    break
        if not matched and by_name.get(_normalize_header(name)):
            warnings.append('기존 원초와 이름은 같은데 code가 다를 수 있습니다.')
    if code and received_date and not lot and matched:
        warnings.append('동일 유사 lot 존재 가능성')

    workplace = _normalize_text(parsed.get('workplace')) or (matched.get('workplace') if matched else '1동 조미')
    if workplace not in _raw_workplace_choices():
        warnings.append('작업장 값이 기본값으로 보정됩니다.')
        workplace = '1동 조미'

    status = 'ERROR' if errors else ('WARNING' if warnings else 'OK')
    return {
        **parsed,
        'target_type': 'raw_material',
        'matched_raw_material_id': int(matched['id']) if matched else None,
        'code': code,
        'name': name,
        'current_quantity': qty,
        'qty': qty,
        'received_date': received_date,
        'lot': lot,
        'ja_ho': ja_ho,
        'sheets_per_sok': sheets_per_sok,
        'car_number': _clean_optional_text(parsed.get('car_number')),
        'workplace': workplace,
        'status': status,
        'error_message': ' | '.join(errors),
        'warning_message': ' | '.join(warnings),
    }


def _row_to_storage_payload(parsed: dict[str, Any]) -> dict[str, Any]:
    keys = [
        'target_type', 'matched_material_id', 'matched_raw_material_id', 'supplier_id',
        'supplier_name', 'code', 'name', 'category', 'spec', 'unit', 'qty',
        'received_quantity', 'current_quantity', 'received_date', 'manufacture_date',
        'expiry_date', 'lot', 'lot_seq', 'supplier_lot', 'ja_ho', 'sheets_per_sok',
        'car_number', 'workplace', 'unit_price', 'status', 'error_message',
        'warning_message',
    ]
    return {key: parsed.get(key) for key in keys}


def _save_parsed_row(cursor, batch_id: int, sheet_name: str, row_no: int, parsed: dict[str, Any]) -> None:
    payload = _row_to_storage_payload(parsed)
    cursor.execute(
        '''
        INSERT INTO import_parsed_rows (
            batch_id, sheet_name, row_no, target_type, matched_material_id, matched_raw_material_id,
            supplier_id, supplier_name, code, name, category, spec, unit, qty,
            received_quantity, current_quantity, received_date, manufacture_date, expiry_date,
            lot, lot_seq, supplier_lot, ja_ho, sheets_per_sok, car_number, workplace,
            unit_price, status, error_message, warning_message
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''',
        (
            batch_id, sheet_name, row_no, payload['target_type'], payload['matched_material_id'], payload['matched_raw_material_id'],
            payload['supplier_id'], payload['supplier_name'], payload['code'], payload['name'], payload['category'], payload['spec'], payload['unit'], payload['qty'],
            payload['received_quantity'], payload['current_quantity'], payload['received_date'], payload['manufacture_date'], payload['expiry_date'],
            payload['lot'], payload['lot_seq'], payload['supplier_lot'], payload['ja_ho'], payload['sheets_per_sok'], payload['car_number'], payload['workplace'],
            payload['unit_price'], payload['status'], payload['error_message'], payload['warning_message'],
        ),
    )

def _recompute_batch_stats(cursor, batch_id: int) -> dict[str, int]:
    row = cursor.execute(
        '''
        SELECT
            COUNT(*) AS total_rows,
            SUM(CASE WHEN status = 'OK' THEN 1 ELSE 0 END) AS ok_rows,
            SUM(CASE WHEN status = 'WARNING' THEN 1 ELSE 0 END) AS warning_rows,
            SUM(CASE WHEN status = 'ERROR' THEN 1 ELSE 0 END) AS error_rows
        FROM import_parsed_rows
        WHERE batch_id = ?
        ''',
        (batch_id,),
    ).fetchone()
    stats = {
        'total_rows': int((row['total_rows'] if row else 0) or 0),
        'ok_rows': int((row['ok_rows'] if row else 0) or 0),
        'warning_rows': int((row['warning_rows'] if row else 0) or 0),
        'error_rows': int((row['error_rows'] if row else 0) or 0),
    }
    cursor.execute(
        '''
        UPDATE import_batches
        SET total_rows = ?, ok_rows = ?, warning_rows = ?, error_rows = ?, status = 'validated'
        WHERE id = ?
        ''',
        (stats['total_rows'], stats['ok_rows'], stats['warning_rows'], stats['error_rows'], batch_id),
    )
    return stats


def _parse_material_dataframe(cursor, batch_id: int, sheet_name: str, df: pd.DataFrame, mapping: dict[str, str], seq_cache: dict[tuple[str, str], int]):
    df = df.where(pd.notna(df), None)
    for idx, row in df.iterrows():
        raw_row = {str(col): (None if pd.isna(val) else val) for col, val in row.items()}
        if not any(_normalize_text(v) for v in raw_row.values()):
            continue
        row_no = int(idx) + 2
        cursor.execute(
            '''
            INSERT INTO import_raw_rows (batch_id, sheet_name, row_no, raw_json)
            VALUES (?, ?, ?, ?)
            ''',
            (batch_id, sheet_name, row_no, _json_dumps(raw_row)),
        )
        parsed_seed = {field: _extract_mapped_value(raw_row, mapping, field) for field in MATERIAL_FIELDS}
        parsed = _validate_material_parsed(cursor, parsed_seed, seq_cache)
        _save_parsed_row(cursor, batch_id, sheet_name, row_no, parsed)


def _parse_raw_dataframe(cursor, batch_id: int, sheet_name: str, df: pd.DataFrame, mapping: dict[str, str]):
    df = df.where(pd.notna(df), None)
    for idx, row in df.iterrows():
        raw_row = {str(col): (None if pd.isna(val) else val) for col, val in row.items()}
        if not any(_normalize_text(v) for v in raw_row.values()):
            continue
        row_no = int(idx) + 2
        cursor.execute(
            '''
            INSERT INTO import_raw_rows (batch_id, sheet_name, row_no, raw_json)
            VALUES (?, ?, ?, ?)
            ''',
            (batch_id, sheet_name, row_no, _json_dumps(raw_row)),
        )
        parsed_seed = {field: _extract_mapped_value(raw_row, mapping, field) for field in RAW_FIELDS}
        parsed = _validate_raw_parsed(cursor, parsed_seed)
        _save_parsed_row(cursor, batch_id, sheet_name, row_no, parsed)


def _parse_batch(conn, batch_id: int, file_path: Path, import_type: str) -> None:
    cursor = conn.cursor()
    seq_cache: dict[tuple[str, str], int] = {}

    cursor.execute('DELETE FROM import_raw_rows WHERE batch_id = ?', (batch_id,))
    cursor.execute('DELETE FROM import_parsed_rows WHERE batch_id = ?', (batch_id,))

    if import_type == 'material_stock' and _parse_material_ledger_workbook(cursor, batch_id, file_path, seq_cache):
        return

    excel = pd.ExcelFile(file_path)
    mapping_summary = []

    for sheet_name in excel.sheet_names:
        df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=object)
        headers = [str(col) for col in df.columns]
        mapping = _detect_column_mapping(headers, import_type)
        mapping_summary.append({'sheet_name': sheet_name, 'mapping': mapping, 'headers': headers})
        if import_type == 'material_stock':
            _parse_material_dataframe(cursor, batch_id, sheet_name, df, mapping, seq_cache)
        elif import_type == 'raw_material_stock':
            _parse_raw_dataframe(cursor, batch_id, sheet_name, df, mapping)
        else:
            raise ValueError(f'지원하지 않는 import_type: {import_type}')

    stats = _recompute_batch_stats(cursor, batch_id)
    cursor.execute(
        '''
        UPDATE import_batches
        SET column_mapping_json = ?, status = 'validated'
        WHERE id = ?
        ''',
        (_json_dumps(mapping_summary), batch_id),
    )
    audit_log(conn, 'validate', 'import_batches', batch_id, stats)


def _get_batch_or_404(cursor, batch_id: int):
    batch = cursor.execute('SELECT * FROM import_batches WHERE id = ?', (batch_id,)).fetchone()
    if not batch:
        raise ValueError('존재하지 않는 임포트 배치입니다.')
    return batch


def _revalidate_one_row(cursor, parsed_row: dict[str, Any]) -> dict[str, Any]:
    seed = dict(parsed_row)
    if seed.get('target_type') == 'material':
        return _validate_material_parsed(cursor, seed, {})
    return _validate_raw_parsed(cursor, seed)


def _update_parsed_row(cursor, row_id: int, parsed: dict[str, Any]) -> None:
    payload = _row_to_storage_payload(parsed)
    cursor.execute(
        '''
        UPDATE import_parsed_rows
        SET target_type = ?, matched_material_id = ?, matched_raw_material_id = ?, supplier_id = ?, supplier_name = ?,
            code = ?, name = ?, category = ?, spec = ?, unit = ?, qty = ?, received_quantity = ?, current_quantity = ?,
            received_date = ?, manufacture_date = ?, expiry_date = ?, lot = ?, lot_seq = ?, supplier_lot = ?, ja_ho = ?,
            sheets_per_sok = ?, car_number = ?, workplace = ?, unit_price = ?, status = ?, error_message = ?, warning_message = ?
        WHERE id = ?
        ''',
        (
            payload['target_type'], payload['matched_material_id'], payload['matched_raw_material_id'], payload['supplier_id'], payload['supplier_name'],
            payload['code'], payload['name'], payload['category'], payload['spec'], payload['unit'], payload['qty'], payload['received_quantity'], payload['current_quantity'],
            payload['received_date'], payload['manufacture_date'], payload['expiry_date'], payload['lot'], payload['lot_seq'], payload['supplier_lot'], payload['ja_ho'],
            payload['sheets_per_sok'], payload['car_number'], payload['workplace'], payload['unit_price'], payload['status'], payload['error_message'], payload['warning_message'],
            row_id,
        ),
    )


def _get_or_create_supplier(cursor, supplier_name: str | None) -> int | None:
    name = _normalize_text(supplier_name)
    if not name:
        return None
    row = cursor.execute('SELECT id FROM suppliers WHERE name = ? LIMIT 1', (name,)).fetchone()
    if row:
        return int(row['id'])
    cursor.execute('INSERT INTO suppliers (name) VALUES (?)', (name,))
    return int(cursor.lastrowid)


def _get_canonical_logistics_location_id(cursor) -> int:
    row = cursor.execute(
        '''
        SELECT id
        FROM inv_locations
        WHERE name = '물류창고'
           OR workplace_code IN ('WH', ?)
           OR (COALESCE(loc_type, '') = 'WAREHOUSE')
        ORDER BY CASE WHEN name = '물류창고' THEN 0
                      WHEN COALESCE(workplace_code, '') = 'WH' THEN 1
                      WHEN COALESCE(workplace_code, '') = ? THEN 2
                      ELSE 3 END,
                 id
        LIMIT 1
        ''',
        (LOGISTICS_WORKPLACE, LOGISTICS_WORKPLACE),
    ).fetchone()
    if row:
        return int(row['id'])
    cursor.execute(
        '''
        INSERT INTO inv_locations (name, loc_type, workplace_code, is_active)
        VALUES ('물류창고', 'WAREHOUSE', 'WH', 1)
        '''
    )
    return int(cursor.lastrowid)


def _get_or_create_location(cursor, workplace: str) -> int:
    name = _normalize_text(workplace) or '1동 조미'
    if name in {LOGISTICS_WORKPLACE, '물류창고'}:
        return _get_canonical_logistics_location_id(cursor)
    row = cursor.execute(
        '''
        SELECT id FROM inv_locations
        WHERE name = ? OR workplace_code = ?
        LIMIT 1
        ''',
        (name, name),
    ).fetchone()
    if row:
        return int(row['id'])
    loc_type = 'logistics' if name == LOGISTICS_WORKPLACE else 'workplace'
    cursor.execute(
        '''
        INSERT INTO inv_locations (name, loc_type, workplace_code, is_active)
        VALUES (?, ?, ?, 1)
        ''',
        (name, loc_type, name),
    )
    return int(cursor.lastrowid)


def _sync_logistics_stock_for_material(cursor, material_id: int, updated_by: str | None = None) -> None:
    material = cursor.execute(
        '''
        SELECT id, code, name, unit
        FROM materials
        WHERE id = ?
        ''',
        (material_id,),
    ).fetchone()
    if not material:
        return
    logistics_location_id = _get_canonical_logistics_location_id(cursor)
    qty_row = cursor.execute(
        '''
        SELECT COALESCE(SUM(b.qty), 0) AS qty
        FROM inv_material_lot_balances b
        JOIN material_lots ml ON ml.id = b.material_lot_id
        WHERE b.location_id = ?
          AND ml.material_id = ?
          AND COALESCE(ml.is_disposed, 0) = 0
        ''',
        (logistics_location_id, material_id),
    ).fetchone()
    logistics_qty = float((qty_row['qty'] if qty_row else 0) or 0)
    material_code = _normalize_text(material['code']) or f"M{int(material['id']):05d}"
    existing = cursor.execute(
        'SELECT id FROM logistics_stocks WHERE material_code = ? LIMIT 1',
        (material_code,),
    ).fetchone()
    if logistics_qty > 0:
        if existing:
            cursor.execute(
                '''
                UPDATE logistics_stocks
                SET material_name = ?, unit = ?, quantity = ?, updated_by = ?, updated_at = CURRENT_TIMESTAMP
                WHERE material_code = ?
                ''',
                (material['name'], material['unit'], logistics_qty, updated_by, material_code),
            )
        else:
            cursor.execute(
                '''
                INSERT INTO logistics_stocks (material_code, material_name, unit, quantity, updated_by)
                VALUES (?, ?, ?, ?, ?)
                ''',
                (material_code, material['name'], material['unit'], logistics_qty, updated_by),
            )
    elif existing:
        cursor.execute('DELETE FROM logistics_stocks WHERE material_code = ?', (material_code,))


def _recalculate_material_stock(cursor, material_id: int) -> None:
    row = cursor.execute(
        '''
        SELECT COALESCE(SUM(current_quantity), 0) AS total_qty
        FROM material_lots
        WHERE material_id = ? AND COALESCE(is_disposed, 0) = 0
        ''',
        (material_id,),
    ).fetchone()
    total_qty = float((row['total_qty'] if row else 0) or 0)
    cursor.execute('UPDATE materials SET current_stock = ? WHERE id = ?', (total_qty, material_id))


def _recalculate_lot_current_quantity(cursor, material_lot_id: int) -> float:
    row = cursor.execute(
        '''
        SELECT COALESCE(SUM(qty), 0) AS total_qty
        FROM inv_material_lot_balances
        WHERE material_lot_id = ?
        ''',
        (material_lot_id,),
    ).fetchone()
    total_qty = float((row['total_qty'] if row else 0) or 0)
    cursor.execute(
        '''
        UPDATE material_lots
        SET current_quantity = ?, quantity = CASE WHEN quantity < ? THEN ? ELSE quantity END, received_quantity = CASE WHEN received_quantity < ? THEN ? ELSE received_quantity END
        WHERE id = ?
        ''',
        (total_qty, total_qty, total_qty, total_qty, total_qty, material_lot_id),
    )
    return total_qty

def _apply_material_row(cursor, batch_id: int, row: dict[str, Any], create_missing: bool, username: str) -> dict[str, int]:
    result = {'materials_created': 0, 'lots_created': 0, 'lots_updated': 0, 'txns_created': 0}
    material_id = row.get('matched_material_id')
    supplier_id = row.get('supplier_id') or _get_or_create_supplier(cursor, row.get('supplier_name'))
    if not material_id:
        if not create_missing:
            raise ValueError('매핑되지 않은 부자재가 있어 자동 생성 옵션 없이 반영할 수 없습니다.')
        cursor.execute(
            '''
            INSERT INTO materials (supplier_id, code, name, category, spec, unit, unit_price, workplace)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''',
            (
                supplier_id,
                _normalize_text(row.get('code')) or None,
                _normalize_text(row.get('name')) or '미지정 부자재',
                _normalize_text(row.get('category')) or _guess_material_category(row.get('name')),
                _normalize_text(row.get('spec')) or None,
                _normalize_text(row.get('unit')) or None,
                float(row.get('unit_price') or 0),
                _normalize_text(row.get('workplace')) or '1동 조미',
            ),
        )
        material_id = int(cursor.lastrowid)
        result['materials_created'] += 1
        if row.get('name'):
            cursor.execute(
                '''
                INSERT OR IGNORE INTO material_name_mappings (source_text, material_id)
                VALUES (?, ?)
                ''',
                (_normalize_text(row['name']), material_id),
            )
    elif supplier_id:
        cursor.execute(
            '''
            UPDATE materials
            SET supplier_id = ?
            WHERE id = ?
            ''',
            (supplier_id, material_id),
        )

    workplace = _normalize_text(row.get('workplace')) or '1동 조미'
    location_id = _get_or_create_location(cursor, workplace)
    lot = _normalize_text(row.get('lot'))
    source_parser = _normalize_text(row.get('source_parser'))
    lot_row = cursor.execute('SELECT * FROM material_lots WHERE lot = ? LIMIT 1', (lot,)).fetchone() if lot else None
    if not lot_row and source_parser == 'material_ledger':
        lot_row = cursor.execute(
            '''
            SELECT ml.*
            FROM material_lots ml
            JOIN inv_material_lot_balances b ON b.material_lot_id = ml.id
            WHERE ml.material_id = ?
              AND COALESCE(ml.receiving_date, '') = COALESCE(?, '')
              AND b.location_id = ?
            ORDER BY ml.id DESC
            LIMIT 1
            ''',
            (material_id, row.get('received_date'), location_id),
        ).fetchone()
        if lot_row:
            lot = lot_row['lot']
    if lot_row and int(lot_row['material_id']) != int(material_id):
        raise ValueError(f'lot 충돌: {lot}')

    if lot_row:
        lot_id = int(lot_row['id'])
        cursor.execute(
            '''
            UPDATE material_lots
            SET receiving_date = ?, manufacture_date = ?, expiry_date = ?, unit_price = ?,
                received_quantity = ?, supplier_lot = ?, lot_seq = ?, quantity = ?
            WHERE id = ?
            ''',
            (
                row.get('received_date'), row.get('manufacture_date'), row.get('expiry_date'), float(row.get('unit_price') or 0),
                float(row.get('received_quantity') or row.get('current_quantity') or 0), _clean_optional_text(row.get('supplier_lot')),
                int(row.get('lot_seq') or 1), float(row.get('received_quantity') or row.get('current_quantity') or 0), lot_id,
            ),
        )
        result['lots_updated'] += 1
    else:
        cursor.execute(
            '''
            INSERT INTO material_lots (
                material_id, lot, receiving_date, manufacture_date, expiry_date,
                unit_price, quantity, lot_seq, received_quantity, current_quantity, supplier_lot
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''',
            (
                material_id, lot, row.get('received_date'), row.get('manufacture_date'), row.get('expiry_date'),
                float(row.get('unit_price') or 0), float(row.get('received_quantity') or row.get('current_quantity') or 0),
                int(row.get('lot_seq') or 1), float(row.get('received_quantity') or row.get('current_quantity') or 0),
                float(row.get('current_quantity') or 0), _clean_optional_text(row.get('supplier_lot')),
            ),
        )
        lot_id = int(cursor.lastrowid)
        result['lots_created'] += 1

    current_qty = float(row.get('current_quantity') or 0)
    balance_row = cursor.execute(
        '''
        SELECT COALESCE(qty, 0) AS qty
        FROM inv_material_lot_balances
        WHERE location_id = ? AND material_lot_id = ?
        ''',
        (location_id, lot_id),
    ).fetchone()
    previous_qty = float((balance_row['qty'] if balance_row else 0) or 0)
    cursor.execute(
        '''
        INSERT INTO inv_material_lot_balances (location_id, material_lot_id, qty, updated_at)
        VALUES (?, ?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(location_id, material_lot_id)
        DO UPDATE SET qty = excluded.qty, updated_at = CURRENT_TIMESTAMP
        ''',
        (location_id, lot_id, current_qty),
    )
    delta = current_qty - previous_qty
    cursor.execute(
        '''
        INSERT INTO inv_material_txns (
            txn_type, location_from_id, location_to_id, material_id, material_lot_id,
            qty, ref_type, ref_id, note, created_by
        ) VALUES ('ADJUST', NULL, ?, ?, ?, ?, 'INITIAL_IMPORT', ?, ?, ?)
        ''',
        (location_id, material_id, lot_id, delta, batch_id, f'초기 엑셀 임포트(batch:{batch_id})', username),
    )
    result['txns_created'] += 1
    _recalculate_lot_current_quantity(cursor, lot_id)
    _recalculate_material_stock(cursor, material_id)
    _sync_logistics_stock_for_material(cursor, material_id, username)
    return result


def _apply_raw_row(cursor, batch_id: int, row: dict[str, Any], username: str) -> dict[str, int]:
    result = {'raw_created': 0, 'raw_updated': 0, 'logs_created': 0}
    raw_id = row.get('matched_raw_material_id')
    current_stock = float(row.get('current_quantity') or 0)
    if raw_id:
        existing = cursor.execute('SELECT * FROM raw_materials WHERE id = ?', (raw_id,)).fetchone()
        if not existing:
            raise ValueError('매핑된 원초를 찾을 수 없습니다.')
        used_quantity = float(existing['used_quantity'] or 0)
        total_stock = max(float(existing['total_stock'] or 0), current_stock + used_quantity)
        cursor.execute(
            '''
            UPDATE raw_materials
            SET name = ?, sheets_per_sok = ?, receiving_date = ?, car_number = ?,
                total_stock = ?, current_stock = ?, code = ?, lot = ?, ja_ho = ?, workplace = ?
            WHERE id = ?
            ''',
            (
                _normalize_text(row.get('name')) or existing['name'],
                int(row.get('sheets_per_sok') or existing['sheets_per_sok'] or 0),
                row.get('received_date') or existing['receiving_date'],
                _clean_optional_text(row.get('car_number')) or existing['car_number'],
                total_stock,
                current_stock,
                _clean_optional_text(row.get('code')) or existing['code'],
                _clean_optional_text(row.get('lot')) or existing['lot'],
                _clean_optional_text(row.get('ja_ho')) or existing['ja_ho'],
                _normalize_text(row.get('workplace')) or existing['workplace'] or '1동 조미',
                raw_id,
            ),
        )
        result['raw_updated'] += 1
    else:
        cursor.execute(
            '''
            INSERT INTO raw_materials (
                name, sheets_per_sok, receiving_date, car_number,
                total_stock, current_stock, used_quantity, workplace, code, lot, ja_ho
            ) VALUES (?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?)
            ''',
            (
                _normalize_text(row.get('name')) or '미지정 원초',
                int(row.get('sheets_per_sok') or 0),
                row.get('received_date'),
                _clean_optional_text(row.get('car_number')),
                current_stock,
                current_stock,
                _normalize_text(row.get('workplace')) or '1동 조미',
                _clean_optional_text(row.get('code')),
                _clean_optional_text(row.get('lot')),
                _clean_optional_text(row.get('ja_ho')),
            ),
        )
        raw_id = int(cursor.lastrowid)
        result['raw_created'] += 1
        if row.get('name'):
            cursor.execute(
                '''
                INSERT OR IGNORE INTO raw_material_name_mappings (source_text, raw_material_code, raw_material_name)
                VALUES (?, ?, ?)
                ''',
                (_normalize_text(row['name']), _normalize_text(row.get('code')), _normalize_text(row.get('name'))),
            )
    cursor.execute(
        '''
        INSERT INTO raw_material_logs (raw_material_id, type, quantity, note, created_by)
        VALUES (?, 'INITIAL_IMPORT', ?, ?, ?)
        ''',
        (raw_id, current_stock, f'초기 엑셀 임포트(batch:{batch_id})', username),
    )
    result['logs_created'] += 1
    return result


def _apply_batch(conn, batch_id: int, create_missing_materials: bool) -> dict[str, Any]:
    cursor = conn.cursor()
    _get_batch_or_404(cursor, batch_id)
    username = (session.get('user') or {}).get('username') or 'system'
    rows = cursor.execute(
        '''
        SELECT *
        FROM import_parsed_rows
        WHERE batch_id = ?
          AND status IN ('OK', 'WARNING')
          AND applied_at IS NULL
        ORDER BY sheet_name, row_no, id
        ''',
        (batch_id,),
    ).fetchall()

    summary = {
        'applied_rows': 0,
        'skipped_rows': 0,
        'materials_created': 0,
        'material_lots_created': 0,
        'material_lots_updated': 0,
        'material_txns_created': 0,
        'raw_materials_created': 0,
        'raw_materials_updated': 0,
        'raw_logs_created': 0,
        'failures': [],
    }

    conn.execute('BEGIN')
    try:
        for row in rows:
            item = dict(row)
            try:
                if item['target_type'] == 'material':
                    result = _apply_material_row(cursor, batch_id, item, create_missing_materials, username)
                    summary['materials_created'] += result['materials_created']
                    summary['material_lots_created'] += result['lots_created']
                    summary['material_lots_updated'] += result['lots_updated']
                    summary['material_txns_created'] += result['txns_created']
                else:
                    result = _apply_raw_row(cursor, batch_id, item, username)
                    summary['raw_materials_created'] += result['raw_created']
                    summary['raw_materials_updated'] += result['raw_updated']
                    summary['raw_logs_created'] += result['logs_created']
                cursor.execute('UPDATE import_parsed_rows SET applied_at = CURRENT_TIMESTAMP WHERE id = ?', (item['id'],))
                summary['applied_rows'] += 1
            except Exception as row_exc:
                summary['skipped_rows'] += 1
                summary['failures'].append({
                    'row_id': item['id'],
                    'sheet_name': item['sheet_name'],
                    'row_no': item['row_no'],
                    'name': item['name'],
                    'message': str(row_exc),
                })
        cursor.execute(
            '''
            UPDATE import_batches
            SET status = ?, applied_at = CURRENT_TIMESTAMP, applied_result_json = ?
            WHERE id = ?
            ''',
            ('applied', _json_dumps(summary), batch_id),
        )
        audit_log(conn, 'apply', 'import_batches', batch_id, summary)
        conn.commit()
    except Exception:
        conn.rollback()
        cursor.execute('UPDATE import_batches SET status = ? WHERE id = ?', ('failed', batch_id))
        conn.commit()
        raise
    return summary

@bp.route('/imports', methods=['GET', 'POST'])
@admin_required
def upload_import():
    conn = get_db()
    cursor = conn.cursor()
    try:
        if request.method == 'POST':
            if not _is_import_runtime_ready():
                flash(_missing_runtime_message(), 'danger')
                return redirect(url_for('imports.upload_import'))
            import_type = (request.form.get('import_type') or '').strip()
            upload = request.files.get('file')
            if import_type not in IMPORT_TYPE_LABELS:
                flash('지원하지 않는 업로드 유형입니다.', 'danger')
                return redirect(url_for('imports.upload_import'))
            if not upload or not upload.filename:
                flash('업로드할 엑셀 파일을 선택해주세요.', 'danger')
                return redirect(url_for('imports.upload_import'))
            if not upload.filename.lower().endswith('.xlsx'):
                flash('.xlsx 파일만 업로드할 수 있습니다.', 'danger')
                return redirect(url_for('imports.upload_import'))

            cursor.execute(
                '''
                INSERT INTO import_batches (file_name, import_type, uploaded_by, status)
                VALUES (?, ?, ?, 'uploaded')
                ''',
                (upload.filename, import_type, (session.get('user') or {}).get('username')),
            )
            batch_id = int(cursor.lastrowid)
            stored_name = _make_batch_upload_name(batch_id, upload.filename)
            file_path = UPLOAD_DIR / stored_name
            upload.save(file_path)
            cursor.execute(
                'UPDATE import_batches SET stored_file_name = ?, status = ? WHERE id = ?',
                (stored_name, 'parsed', batch_id),
            )
            _parse_batch(conn, batch_id, file_path, import_type)
            conn.commit()
            flash('엑셀 업로드와 검증이 완료되었습니다. 미리보기에서 확인 후 반영해주세요.', 'success')
            return redirect(url_for('imports.batch_preview', batch_id=batch_id))

        recent_batches = cursor.execute(
            '''
            SELECT *
            FROM import_batches
            ORDER BY id DESC
            LIMIT 10
            '''
        ).fetchall()
        return render_template(
            'imports_upload.html',
            user=session['user'],
            import_type_labels=IMPORT_TYPE_LABELS,
            recent_batches=recent_batches,
            import_runtime_ready=_is_import_runtime_ready(),
            import_runtime_message=_missing_runtime_message() if not _is_import_runtime_ready() else '',
        )
    finally:
        conn.close()


@bp.route('/imports/templates/<template_key>.xlsx')
@admin_required
def download_import_template(template_key):
    try:
        return _build_template_response(template_key)
    except KeyError:
        flash('지원하지 않는 템플릿입니다.', 'danger')
        return redirect(url_for('imports.upload_import'))


@bp.route('/imports/history')
@admin_required
def import_history():
    conn = get_db()
    cursor = conn.cursor()
    try:
        batches = cursor.execute(
            '''
            SELECT *
            FROM import_batches
            ORDER BY id DESC
            LIMIT 200
            '''
        ).fetchall()
        return render_template('imports_history.html', user=session['user'], batches=batches, import_type_labels=IMPORT_TYPE_LABELS)
    finally:
        conn.close()


@bp.route('/imports/<int:batch_id>')
@admin_required
def batch_preview(batch_id: int):
    conn = get_db()
    cursor = conn.cursor()
    try:
        batch = _get_batch_or_404(cursor, batch_id)
        workplace_filter = (request.args.get('workplace') or '전체').strip() or '전체'
        where_sql = 'WHERE batch_id = ?'
        params: list[Any] = [batch_id]
        if workplace_filter != '전체':
            where_sql += ' AND COALESCE(workplace, "") = ?'
            params.append(workplace_filter)
        rows = cursor.execute(
            f'''
            SELECT *
            FROM import_parsed_rows
            {where_sql}
            ORDER BY CASE status WHEN 'ERROR' THEN 0 WHEN 'WARNING' THEN 1 ELSE 2 END, sheet_name, row_no, id
            ''',
            params,
        ).fetchall()
        workplace_rows = cursor.execute(
            '''
            SELECT workplace, COUNT(*) as row_count
            FROM import_parsed_rows
            WHERE batch_id = ? AND COALESCE(workplace, '') <> ''
            GROUP BY workplace
            ORDER BY workplace
            ''',
            (batch_id,),
        ).fetchall()
        workplace_tabs = [{'label': '전체', 'count': int(batch['total_rows'] or 0)}]
        workplace_tabs.extend(
            {'label': row['workplace'], 'count': int(row['row_count'] or 0)}
            for row in workplace_rows
        )
        material_options = _load_material_options(cursor)
        raw_material_options = _load_raw_material_options(cursor)
        mapping_summary = _json_loads(batch['column_mapping_json'], [])
        result_summary = _json_loads(batch['applied_result_json'], {})
        suppliers = cursor.execute('SELECT id, name FROM suppliers ORDER BY name').fetchall()
        return render_template(
            'imports_preview.html',
            user=session['user'],
            batch=batch,
            rows=rows,
            workplace_filter=workplace_filter,
            workplace_tabs=workplace_tabs,
            mapping_summary=mapping_summary,
            import_type_labels=IMPORT_TYPE_LABELS,
            material_options=material_options,
            raw_material_options=raw_material_options,
            suppliers=suppliers,
            material_workplaces=_material_workplace_choices(),
            raw_workplaces=_raw_workplace_choices(),
            result_summary=result_summary,
        )
    finally:
        conn.close()


@bp.route('/imports/<int:batch_id>/rows/<int:row_id>/remap', methods=['POST'])
@admin_required
def remap_batch_row(batch_id: int, row_id: int):
    conn = get_db()
    cursor = conn.cursor()
    try:
        _get_batch_or_404(cursor, batch_id)
        row = cursor.execute('SELECT * FROM import_parsed_rows WHERE id = ? AND batch_id = ?', (row_id, batch_id)).fetchone()
        if not row:
            flash('수정할 임포트 행을 찾을 수 없습니다.', 'danger')
            return redirect(url_for('imports.batch_preview', batch_id=batch_id))
        item = dict(row)
        if item['target_type'] == 'material':
            item['matched_material_id'] = request.form.get('matched_material_id') or None
            item['supplier_id'] = request.form.get('supplier_id') or None
            for field in ['code', 'name', 'category', 'spec', 'unit', 'lot', 'supplier_lot', 'received_date', 'manufacture_date', 'expiry_date', 'workplace', 'supplier_name']:
                if field in request.form:
                    item[field] = request.form.get(field)
            for field in ['qty', 'current_quantity', 'received_quantity', 'unit_price', 'lot_seq']:
                if field in request.form:
                    item[field] = request.form.get(field)
        else:
            item['matched_raw_material_id'] = request.form.get('matched_raw_material_id') or None
            for field in ['code', 'name', 'lot', 'ja_ho', 'received_date', 'car_number', 'workplace']:
                if field in request.form:
                    item[field] = request.form.get(field)
            for field in ['qty', 'current_quantity', 'sheets_per_sok']:
                if field in request.form:
                    item[field] = request.form.get(field)
        validated = _revalidate_one_row(cursor, item)
        _update_parsed_row(cursor, row_id, validated)
        _recompute_batch_stats(cursor, batch_id)
        conn.commit()
        flash('행 매핑을 저장하고 다시 검증했습니다.', 'success')
        return redirect(url_for('imports.batch_preview', batch_id=batch_id))
    finally:
        conn.close()


@bp.route('/imports/<int:batch_id>/revalidate', methods=['POST'])
@admin_required
def revalidate_batch(batch_id: int):
    conn = get_db()
    cursor = conn.cursor()
    try:
        if not _is_import_runtime_ready():
            flash(_missing_runtime_message(), 'danger')
            return redirect(url_for('imports.batch_preview', batch_id=batch_id))
        batch = _get_batch_or_404(cursor, batch_id)
        if not batch['stored_file_name']:
            flash('원본 업로드 파일 정보가 없습니다.', 'danger')
            return redirect(url_for('imports.batch_preview', batch_id=batch_id))
        file_path = UPLOAD_DIR / batch['stored_file_name']
        if not file_path.exists():
            flash('원본 업로드 파일을 찾을 수 없습니다.', 'danger')
            return redirect(url_for('imports.batch_preview', batch_id=batch_id))
        _parse_batch(conn, batch_id, file_path, batch['import_type'])
        conn.commit()
        flash('배치를 다시 파싱하고 재검증했습니다.', 'success')
        return redirect(url_for('imports.batch_preview', batch_id=batch_id))
    finally:
        conn.close()


@bp.route('/imports/<int:batch_id>/apply', methods=['POST'])
@admin_required
def apply_batch(batch_id: int):
    conn = get_db()
    try:
        create_missing_materials = bool(request.form.get('create_missing_materials'))
        _apply_batch(conn, batch_id, create_missing_materials)
        flash('임포트 반영이 완료되었습니다.', 'success')
        return redirect(url_for('imports.batch_result', batch_id=batch_id))
    except Exception as exc:
        flash(f'임포트 반영 중 오류가 발생했습니다: {exc}', 'danger')
        return redirect(url_for('imports.batch_preview', batch_id=batch_id))
    finally:
        conn.close()


@bp.route('/imports/<int:batch_id>/result')
@admin_required
def batch_result(batch_id: int):
    conn = get_db()
    cursor = conn.cursor()
    try:
        batch = _get_batch_or_404(cursor, batch_id)
        result_summary = _json_loads(batch['applied_result_json'], {})
        return render_template(
            'imports_result.html',
            user=session['user'],
            batch=batch,
            result_summary=result_summary,
            import_type_labels=IMPORT_TYPE_LABELS,
        )
    finally:
        conn.close()


@bp.route('/imports/<int:batch_id>/failed.csv')
@admin_required
def download_failed_rows(batch_id: int):
    conn = get_db()
    cursor = conn.cursor()
    try:
        _get_batch_or_404(cursor, batch_id)
        rows = cursor.execute(
            '''
            SELECT sheet_name, row_no, target_type, code, name, status, error_message, warning_message
            FROM import_parsed_rows
            WHERE batch_id = ? AND status = 'ERROR'
            ORDER BY sheet_name, row_no, id
            ''',
            (batch_id,),
        ).fetchall()
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['sheet_name', 'row_no', 'target_type', 'code', 'name', 'status', 'error_message', 'warning_message'])
        for row in rows:
            writer.writerow([row['sheet_name'], row['row_no'], row['target_type'], row['code'], row['name'], row['status'], row['error_message'], row['warning_message']])
        return Response(
            output.getvalue(),
            mimetype='text/csv; charset=utf-8-sig',
            headers={'Content-Disposition': f'attachment; filename=import_batch_{batch_id}_failed_rows.csv'},
        )
    finally:
        conn.close()


@bp.route('/imports/<int:batch_id>/delete', methods=['POST'])
@admin_required
def delete_batch(batch_id: int):
    conn = get_db()
    cursor = conn.cursor()
    try:
        batch = _get_batch_or_404(cursor, batch_id)
        stored_file_name = _normalize_text(batch['stored_file_name'])
        conn.execute('BEGIN')
        cursor.execute('DELETE FROM import_raw_rows WHERE batch_id = ?', (batch_id,))
        cursor.execute('DELETE FROM import_parsed_rows WHERE batch_id = ?', (batch_id,))
        cursor.execute('DELETE FROM import_batches WHERE id = ?', (batch_id,))
        audit_log(
            conn,
            'delete',
            'import_batches',
            batch_id,
            {
                'file_name': batch['file_name'],
                'import_type': batch['import_type'],
                'status': batch['status'],
                'note': '임포트 이력/staging 삭제. 실데이터 롤백 없음.',
            },
        )
        conn.commit()
        if stored_file_name:
            file_path = UPLOAD_DIR / stored_file_name
            if file_path.exists():
                try:
                    file_path.unlink()
                except Exception:
                    pass
        flash('임포트 배치를 삭제했습니다. 실데이터는 되돌리지 않았습니다.', 'success')
    except Exception as exc:
        conn.rollback()
        flash(f'임포트 배치 삭제 중 오류가 발생했습니다: {exc}', 'danger')
    finally:
        conn.close()
    return redirect(request.form.get('next') or url_for('imports.import_history'))
